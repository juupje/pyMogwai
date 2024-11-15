import ast
import os

import pandas as pd
from openpyxl import load_workbook

from mogwai.core import MogwaiGraph

from .filesystem import get_file_stats


class EXCELGraph(MogwaiGraph):
    LABEL = "EXCELFile"

    def __init__(self, file: str = None, name: str = None, **kwargs):
        super().__init__(**kwargs)
        if file is not None:
            if not os.path.exists(file):
                raise FileNotFoundError(f"No such file or directory: '{file:s}'")
            self.file = file
            name = name or os.path.basename(file)
            self.root = self.add_labeled_node(
                label=EXCELGraph.LABEL, name=name, **get_file_stats(self.file)
            )
            self.construct()

    def construct(self):
        dic = excel_to_dic(self.file)
        sheets = dic.pop("sheets")
        self.nodes[self.root].update({"metadata": dic, "number_of_sheets": len(sheets)})
        for sheet in sheets:
            node = self.add_labeled_node(
                label="EXCELSheet", name=sheet.pop("name"), **sheet
            )
            self.add_labeled_edge(self.root, node, "HAS_SHEET")


def excel_to_dic(path: str) -> dict:
    workbook = load_workbook(path)
    p = workbook.properties
    sheetnames = workbook.sheetnames
    dic = {
        "creator": p.creator,
        "title": p.title,
        "description": p.description,
        "subject": p.subject,
        "identifier": p.identifier,
        "language": p.language,
        "lastModifiedBy": p.lastModifiedBy,
        "category": p.category,
        "status": p.contentStatus,
        "version": p.version,
        "revision": p.revision,
        "keywords": p.keywords,
    }
    dic["sheets"] = []

    for sheet in sheetnames:
        try:
            dataframe = pd.read_excel(path, sheet_name=sheet)

            json = ast.literal_eval(pd.DataFrame.to_json(dataframe))
            json = dict({"name": sheet}, **json)  # json["name"] = sheet
            dic["sheets"].append(json)
        except ValueError:
            raise ValueError(f"Excel sheet with path {path} needs a tabular format")
    return dic
